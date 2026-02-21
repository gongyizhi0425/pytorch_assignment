import argparse
import csv
import gc
import json
import os
import platform
import socket
import time
from datetime import datetime, timezone
from dataclasses import asdict, dataclass
from threading import Event, Thread
from typing import Any, Dict, Iterable, List, Optional, Tuple

import psutil
import torch
import torchvision.models as models

try:
    from thop import profile as thop_profile  # type: ignore
except Exception:
    thop_profile = None

try:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
except Exception:
    plt = None

try:
    import yaml  # type: ignore
except Exception:
    yaml = None

try:
    import openpyxl  # type: ignore
    from openpyxl import Workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except Exception:
    openpyxl = None
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore

def _mb(nbytes: int) -> float:
    return nbytes / (1024**2)


def _dtype_from_str(s: str) -> torch.dtype:
    s = s.lower().strip()
    if s in ("fp32", "float32"):
        return torch.float32
    if s in ("fp16", "float16"):
        return torch.float16
    if s in ("bf16", "bfloat16"):
        return torch.bfloat16
    raise ValueError(f"Unknown dtype: {s}")


def _normalize_device(device: str) -> torch.device:
    if device == "auto":
        device = "cuda" if torch.cuda.is_available() else "cpu"
    return torch.device(device)


def _safe_infer_dtype(device: torch.device, dtype_str: str) -> Tuple[torch.dtype, str]:
    dtype = _dtype_from_str(dtype_str)
    if device.type == "cpu" and dtype != torch.float32:
        # CPU 上很多 op 不支持 fp16/bf16，统一退回 fp32。
        return torch.float32, "fp32"
    return dtype, dtype_str


def _module_params_and_bytes(model: torch.nn.Module) -> Tuple[int, int]:
    num_params = sum(p.numel() for p in model.parameters())
    param_bytes = sum(p.numel() * p.element_size() for p in model.parameters())
    buffer_bytes = sum(b.numel() * b.element_size() for b in model.buffers())
    return num_params, (param_bytes + buffer_bytes)


def _module_params_and_numel_including_buffers(model: torch.nn.Module) -> Tuple[int, int]:
    num_params = sum(p.numel() for p in model.parameters())
    total_numel = num_params + sum(b.numel() for b in model.buffers())
    return num_params, total_numel


def _load_torchvision_model(model_name: str, pretrained: bool) -> torch.nn.Module:
    name = model_name.lower().strip()

    # 兼容 torchvision 新旧 weights API
    def _resnet18():
        try:
            weights = models.ResNet18_Weights.DEFAULT if pretrained else None
            return models.resnet18(weights=weights)
        except Exception:
            return models.resnet18(pretrained=pretrained)

    def _mbv3_small():
        try:
            weights = models.MobileNet_V3_Small_Weights.DEFAULT if pretrained else None
            return models.mobilenet_v3_small(weights=weights)
        except Exception:
            return models.mobilenet_v3_small(pretrained=pretrained)

    def _mbv3_large():
        try:
            weights = models.MobileNet_V3_Large_Weights.DEFAULT if pretrained else None
            return models.mobilenet_v3_large(weights=weights)
        except Exception:
            return models.mobilenet_v3_large(pretrained=pretrained)

    if name == "resnet18":
        return _resnet18()
    if name in ("mobilenetv3", "mobilenet_v3_small", "mobilenetv3small"):
        return _mbv3_small()
    if name in ("mobilenet_v3_large", "mobilenetv3large"):
        return _mbv3_large()
    raise ValueError(f"Unknown model: {model_name}")


def _read_config(path: str) -> Dict[str, Any]:
    """Read a config file.
    - Prefer YAML if PyYAML is installed.
    - Otherwise accept JSON (note: JSON is valid YAML 1.2) so users can still use .yaml.
    """
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read()

    if yaml is not None:
        data = yaml.safe_load(raw)
        if data is None:
            return {}
        if not isinstance(data, dict):
            raise ValueError(f"Config root must be a mapping/dict: {path}")
        return dict(data)

    # Fallback: JSON
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise RuntimeError(
            "未安装 PyYAML，且该配置文件不是有效 JSON。\n"
            "解决方案：\n"
            "- 安装 PyYAML：pip install pyyaml\n"
            "- 或把 configs/test_config.yaml 写成 JSON 语法（JSON 也是 YAML 1.2 的子集）\n"
            f"JSON 解析错误：{e}"
        ) from e
    if not isinstance(data, dict):
        raise ValueError(f"Config root must be a JSON object/dict: {path}")
    return dict(data)


def _apply_config_over_args(args: argparse.Namespace, defaults: argparse.Namespace, cfg: Dict[str, Any]) -> None:
    """Apply config values to argparse args unless user explicitly set the CLI flag.

    We detect explicit CLI flags by comparing to parser defaults.
    """

    def _maybe_set(attr: str, value: Any) -> None:
        if not hasattr(args, attr):
            return
        if getattr(args, attr) != getattr(defaults, attr):
            return  # CLI explicitly set
        setattr(args, attr, value)

    # allow both model/models for convenience
    if "model" in cfg and "models" not in cfg:
        _maybe_set("models", str(cfg["model"]))
    if "models" in cfg:
        _maybe_set("models", str(cfg["models"]))

    for k in ("device", "dtype", "warmup", "iters", "out_prefix", "isolate_runs", "no_plot"):
        if k in cfg:
            _maybe_set(k.replace("-", "_"), cfg[k])

    if "pretrained" in cfg:
        # argparse uses store_true; accept bool in config
        _maybe_set("pretrained", bool(cfg["pretrained"]))

    # batch sizes
    if "batch_sizes" in cfg and "batch-sizes" not in cfg:
        bs = cfg["batch_sizes"]
        if isinstance(bs, list):
            _maybe_set("batch_sizes", ",".join(str(int(x)) for x in bs))
        else:
            _maybe_set("batch_sizes", str(bs))
    if "batch-sizes" in cfg:
        _maybe_set("batch_sizes", str(cfg["batch-sizes"]))

    # resolution(s)
    if "resolution" in cfg and "resolutions" not in cfg:
        _maybe_set("resolutions", str(int(cfg["resolution"])))
    if "resolutions" in cfg:
        rs = cfg["resolutions"]
        if isinstance(rs, list):
            _maybe_set("resolutions", ",".join(str(int(x)) for x in rs))
        else:
            _maybe_set("resolutions", str(rs))


def _env_meta(device: torch.device, dtype_str: str, pretrained: bool) -> Dict[str, Any]:
    meta: Dict[str, Any] = {}
    meta["timestamp_utc"] = datetime.now(timezone.utc).isoformat()
    meta["host"] = socket.gethostname()
    meta["platform"] = platform.platform()
    meta["cwd"] = os.getcwd()
    meta["pid"] = os.getpid()
    meta["python"] = platform.python_version()
    meta["torch"] = getattr(torch, "__version__", "unknown")
    meta["torchvision"] = getattr(models, "__package__", "torchvision")
    meta["device"] = device.type
    if device.type == "cuda":
        try:
            meta["cuda_device_name"] = torch.cuda.get_device_name(0)
            meta["cuda_capability"] = ".".join(map(str, torch.cuda.get_device_capability(0)))
        except Exception:
            meta["cuda_device_name"] = "unknown"
    meta["dtype"] = dtype_str
    meta["pretrained"] = bool(pretrained)
    meta["thop_available"] = thop_profile is not None
    meta["matplotlib_available"] = plt is not None
    return meta


def _write_report_csv(
    out_prefix: str,
    meta: Dict[str, Any],
    static_row: "StaticRow",
    runtime_rows: List["RuntimeRow"],
) -> str:
    """Write a single Excel-friendly report CSV (one row per runtime measurement)."""
    rows: List[Dict[str, Any]] = []
    for rr in runtime_rows:
        row: Dict[str, Any] = {}
        row.update(meta)
        row["model"] = rr.model
        row["num_params"] = static_row.num_params
        row["weights_and_buffers_mb"] = round(static_row.weights_and_buffers_mb, 6)
        row["batch_size"] = rr.batch_size
        row["resolution"] = rr.resolution
        row["gmacs_per_image"] = rr.gmacs_per_image
        row["gflops_per_image"] = rr.gflops_per_image
        row["gmacs_per_batch"] = None if rr.gmacs_per_image is None else float(rr.gmacs_per_image) * int(rr.batch_size)
        row["gflops_per_batch"] = None if rr.gflops_per_image is None else float(rr.gflops_per_image) * int(rr.batch_size)
        row["latency_ms_per_image"] = float(rr.latency_ms_per_image)
        row["throughput_images_s"] = float(rr.throughput_images_s)
        row["peak_cpu_rss_mb"] = round(rr.peak_cpu_rss_mb, 6)
        row["peak_gpu_allocated_mb"] = rr.peak_gpu_allocated_mb
        row["activation_total_per_fwd_mb"] = round(rr.activation_total_per_fwd_mb, 6)
        row["activation_max_single_mb"] = round(rr.activation_max_single_mb, 6)
        rows.append(row)

    out_path = f"{out_prefix}_report.csv"
    _write_csv(out_path, rows)
    return out_path


def _write_xlsx(path: str, sheets: Dict[str, List[List[Any]]]) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl 不可用，无法写入 xlsx")

    wb = Workbook()
    # remove the default empty sheet
    wb.remove(wb.active)

    for sheet_name, table in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in table:
            ws.append(row)

        # basic usability: freeze header
        if len(table) >= 2:
            ws.freeze_panes = "A2"

        # auto-ish column width (bounded)
        if get_column_letter is not None and table:
            ncols = max(len(r) for r in table)
            for col_idx in range(1, ncols + 1):
                max_len = 0
                for r in table[: min(len(table), 200)]:
                    if col_idx - 1 >= len(r):
                        continue
                    v = r[col_idx - 1]
                    if v is None:
                        continue
                    max_len = max(max_len, len(str(v)))
                ws.column_dimensions[get_column_letter(col_idx)].width = float(min(max(max_len + 2, 10), 48))

    wb.save(path)


def _write_report_xlsx(
    out_prefix: str,
    meta: Dict[str, Any],
    static_row: "StaticRow",
    runtime_rows: List["RuntimeRow"],
) -> str:
    """Write a single Excel workbook with multiple sheets.

    Sheets:
    - meta: environment + run configuration
    - static: static metrics
    - runtime: runtime rows (one per batch/res)
    """

    # meta sheet (two columns)
    meta_table: List[List[Any]] = [["key", "value"]]
    for k in sorted(meta.keys()):
        meta_table.append([k, meta[k]])

    static_table: List[List[Any]] = [
        ["model", "dtype", "num_params", "weights_and_buffers_mb"],
        [static_row.model, static_row.dtype, int(static_row.num_params), float(static_row.weights_and_buffers_mb)],
    ]

    runtime_header = [
        "model",
        "device",
        "dtype",
        "batch_size",
        "resolution",
        "gmacs_per_image",
        "gflops_per_image",
        "gmacs_per_batch",
        "gflops_per_batch",
        "latency_ms_per_image",
        "throughput_images_s",
        "peak_cpu_rss_mb",
        "peak_gpu_allocated_mb",
        "activation_total_per_fwd_mb",
        "activation_max_single_mb",
    ]
    runtime_table: List[List[Any]] = [runtime_header]
    for r in runtime_rows:
        gmacs_per_batch = None if r.gmacs_per_image is None else float(r.gmacs_per_image) * int(r.batch_size)
        gflops_per_batch = None if r.gflops_per_image is None else float(r.gflops_per_image) * int(r.batch_size)
        runtime_table.append(
            [
                r.model,
                r.device,
                r.dtype,
                int(r.batch_size),
                int(r.resolution),
                r.gmacs_per_image,
                r.gflops_per_image,
                gmacs_per_batch,
                gflops_per_batch,
                float(r.latency_ms_per_image),
                float(r.throughput_images_s),
                float(r.peak_cpu_rss_mb),
                r.peak_gpu_allocated_mb,
                float(r.activation_total_per_fwd_mb),
                float(r.activation_max_single_mb),
            ]
        )

    out_path = f"{out_prefix}_report.xlsx"
    _write_xlsx(out_path, {"meta": meta_table, "static": static_table, "runtime": runtime_table})
    return out_path


def _write_detailed_report(
    out_prefix: str,
    meta: Dict[str, Any],
    static_row: "StaticRow",
    runtime_rows: List["RuntimeRow"],
) -> str:
    """Prefer XLSX; fallback to CSV if xlsx writer is unavailable."""
    if openpyxl is not None:
        return _write_report_xlsx(out_prefix, meta, static_row, runtime_rows)
    print("[note] 未安装 openpyxl，详细报告将退回 CSV（Excel 仍可打开）。可选安装：pip install openpyxl")
    return _write_report_csv(out_prefix, meta, static_row, runtime_rows)

def _rss_mb(proc: psutil.Process) -> float:
    return proc.memory_info().rss / (1024**2)


def _peak_rss_during(fn, poll_s: float = 0.005) -> float:
    """在当前进程内，用采样方式估计 fn 运行期间的 peak RSS (MB)。"""
    proc = psutil.Process()
    stop = Event()
    peak = {"v": _rss_mb(proc)}

    def _poll() -> None:
        while not stop.is_set():
            peak["v"] = max(peak["v"], _rss_mb(proc))
            time.sleep(poll_s)

    t = Thread(target=_poll, daemon=True)
    t.start()
    try:
        fn()
    finally:
        stop.set()
        t.join(timeout=1)

    return float(peak["v"])


class ActivationProxy:
    """CNN 激活内存 proxy：用 forward hook 统计输出张量大小。

    - total_bytes_this_fwd：本次前向中所有 layer 输出张量大小之和（粗略 proxy）
    - max_single_tensor_bytes_this_fwd：本次前向中单个输出张量最大值
    """

    def __init__(self) -> None:
        self.total_bytes_this_fwd = 0
        self.max_single_tensor_bytes_this_fwd = 0
        self._handles: List[torch.utils.hooks.RemovableHandle] = []

    @staticmethod
    def _tensor_bytes(t: torch.Tensor) -> int:
        return t.numel() * t.element_size()

    def _hook(self, _module: torch.nn.Module, _inp, out) -> None:
        def visit(x) -> None:
            if torch.is_tensor(x):
                b = self._tensor_bytes(x)
                self.total_bytes_this_fwd += b
                self.max_single_tensor_bytes_this_fwd = max(self.max_single_tensor_bytes_this_fwd, b)
            elif isinstance(x, (tuple, list)):
                for y in x:
                    visit(y)
            elif isinstance(x, dict):
                for y in x.values():
                    visit(y)

        visit(out)

    def reset(self) -> None:
        self.total_bytes_this_fwd = 0
        self.max_single_tensor_bytes_this_fwd = 0

    def install(self, model: torch.nn.Module) -> None:
        self.remove()
        for m in model.modules():
            # 只在叶子模块上挂钩子，避免重复统计
            if len(list(m.children())) > 0:
                continue
            self._handles.append(m.register_forward_hook(self._hook))

    def remove(self) -> None:
        for h in self._handles:
            try:
                h.remove()
            except Exception:
                pass
        self._handles = []

@dataclass
class StaticRow:
    model: str
    dtype: str
    num_params: int
    weights_and_buffers_mb: float


@dataclass
class RuntimeRow:
    model: str
    device: str
    dtype: str
    batch_size: int
    resolution: int
    gmacs_per_image: Optional[float]
    gflops_per_image: Optional[float]
    latency_ms_per_image: float
    throughput_images_s: float
    peak_cpu_rss_mb: float
    peak_gpu_allocated_mb: Optional[float]
    activation_total_per_fwd_mb: float
    activation_max_single_mb: float


def _compute_macs_flops_per_image(
    model_name: str,
    resolution: int,
    pretrained: bool,
) -> Tuple[Optional[float], Optional[float]]:
    """返回 (GMACs/img, GFLOPs/img)。

    说明：
    - 依赖可选包 thop；未安装则返回 (None, None)
    - FLOPs 这里按常见约定：1 MAC = 1 multiply-add，FLOPs ~= 2 * MACs
    - MACs/FLOPs 与权重是否 pretrained 无关，但这里仍接收 pretrained 以便复用同一构建路径。
    """
    if thop_profile is None:
        return None, None

    model = _load_torchvision_model(model_name, pretrained=pretrained).eval()
    x = torch.randn(1, 3, resolution, resolution)
    with torch.inference_mode():
        macs, _params = thop_profile(model, inputs=(x,), verbose=False)

    # thop 返回的是 MACs（乘加次数），这里换算 FLOPs
    gmacs = float(macs) / 1e9
    gflops = float(macs) * 2.0 / 1e9

    del model
    gc.collect()
    return gmacs, gflops


def _run_inference_sweep(
    model: torch.nn.Module,
    device: torch.device,
    dtype_str: str,
    batch_size: int,
    resolution: int,
    warmup: int,
    iters: int,
) -> Tuple[float, Optional[float], float, float, float, float]:
    """返回：peak_cpu_rss_mb, peak_gpu_allocated_mb, act_total_mb, act_max_mb, latency_ms_per_image, throughput_images_s"""
    dtype, dtype_str2 = _safe_infer_dtype(device, dtype_str)
    model = model.to(device=device, dtype=dtype).eval()
    x = torch.randn(batch_size, 3, resolution, resolution, device=device, dtype=dtype)

    act = ActivationProxy()
    act.install(model)

    # warmup（不计入 peak）
    with torch.inference_mode():
        for _ in range(warmup):
            act.reset()
            _ = model(x)
            if device.type == "cuda":
                torch.cuda.synchronize()

    # measured iters：在这里 reset GPU peak，并用采样线程捕获 CPU peak RSS
    if device.type == "cuda":
        torch.cuda.synchronize()
        torch.cuda.reset_peak_memory_stats()

    timings_s: List[float] = []
    act_total_max = 0
    act_single_max = 0

    def _measured_loops() -> None:
        nonlocal act_total_max, act_single_max
        with torch.inference_mode():
            for _ in range(max(1, iters)):
                act.reset()
                # make CUDA timings accurate
                if device.type == "cuda":
                    torch.cuda.synchronize()
                t0 = time.perf_counter()
                _ = model(x)
                if device.type == "cuda":
                    torch.cuda.synchronize()
                t1 = time.perf_counter()
                timings_s.append(t1 - t0)
                act_total_max = max(act_total_max, act.total_bytes_this_fwd)
                act_single_max = max(act_single_max, act.max_single_tensor_bytes_this_fwd)

    peak_rss_mb = _peak_rss_during(_measured_loops)

    peak_gpu_mb = None
    if device.type == "cuda":
        peak_gpu_mb = _mb(int(torch.cuda.max_memory_allocated()))

    act.remove()
    total_time_s = float(sum(timings_s))
    total_images = max(1, int(batch_size) * max(1, int(iters)))
    latency_ms_per_image = float("nan")
    throughput_images_s = float("nan")
    if total_time_s > 0:
        latency_ms_per_image = (total_time_s / total_images) * 1e3
        throughput_images_s = total_images / total_time_s

    return (
        peak_rss_mb,
        peak_gpu_mb,
        _mb(act_total_max),
        _mb(act_single_max),
        latency_ms_per_image,
        throughput_images_s,
    )


def _parse_int_list(s: str) -> List[int]:
    """Parse comma-separated ints and ranges.

    Examples:
    - "1,4,16" -> [1,4,16]
    - "1-16" -> [1..16]
    - "1-4,8,16" -> [1,2,3,4,8,16]
    """
    items: List[int] = []
    for part in [p.strip() for p in s.split(",") if p.strip()]:
        if "-" in part:
            a, b = part.split("-", 1)
            start = int(a.strip())
            end = int(b.strip())
            step = 1 if end >= start else -1
            items.extend(list(range(start, end + step, step)))
        else:
            items.append(int(part))
    # de-dup while keeping order
    seen = set()
    out: List[int] = []
    for x in items:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _group_by(items: Iterable[RuntimeRow], key_fn) -> Dict:
    out: Dict = {}
    for it in items:
        k = key_fn(it)
        out.setdefault(k, []).append(it)
    return out


def _plot_runtime(runtime_rows: List[RuntimeRow], out_prefix: str) -> None:
    """Save comparison plots as PNG. No GUI required."""
    if plt is None:
        print("[note] 未安装 matplotlib，跳过绘图。可选安装：pip install matplotlib")
        return
    if not runtime_rows:
        return

    # Sort for stable lines
    runtime_rows = sorted(runtime_rows, key=lambda r: (r.resolution, r.model, r.batch_size))
    models_sorted = sorted({r.model for r in runtime_rows})
    resolutions_sorted = sorted({r.resolution for r in runtime_rows})

    def plot_metric(
        metric_name: str,
        y_label: str,
        y_getter,
        filename: str,
        skip_all_na: bool = True,
    ) -> None:
        n = len(resolutions_sorted)
        fig, axes = plt.subplots(1, n, figsize=(5.5 * n, 4.2), squeeze=False)
        any_data = False
        for i, res in enumerate(resolutions_sorted):
            ax = axes[0][i]
            ax.set_title(f"res={res}")
            ax.set_xlabel("batch size")
            ax.set_ylabel(y_label)
            ax.grid(True, alpha=0.3)

            plotted_any = False
            for m in models_sorted:
                pts = [r for r in runtime_rows if r.resolution == res and r.model == m]
                pts = sorted(pts, key=lambda r: r.batch_size)
                xs = [r.batch_size for r in pts]
                ys = [y_getter(r) for r in pts]
                # drop None values
                xs2, ys2 = [], []
                for x, y in zip(xs, ys):
                    if y is None:
                        continue
                    xs2.append(x)
                    ys2.append(float(y))
                if xs2:
                    any_data = True
                    plotted_any = True
                    ax.plot(xs2, ys2, marker="o", linewidth=1.5, label=m)

            if plotted_any:
                ax.legend(fontsize=8)

        if skip_all_na and not any_data:
            plt.close(fig)
            return
        fig.suptitle(metric_name)
        fig.tight_layout()
        out_path = f"{out_prefix}_{filename}.png"
        fig.savefig(out_path, dpi=200)
        plt.close(fig)
        print(f"Saved plot: {out_path}")

    plot_metric(
        metric_name="Peak CPU RSS vs batch size",
        y_label="MB",
        y_getter=lambda r: r.peak_cpu_rss_mb,
        filename="peak_rss",
    )
    plot_metric(
        metric_name="Peak GPU allocated vs batch size",
        y_label="MB",
        y_getter=lambda r: r.peak_gpu_allocated_mb,
        filename="peak_gpu",
    )
    plot_metric(
        metric_name="Activation total (proxy) vs batch size",
        y_label="MB",
        y_getter=lambda r: r.activation_total_per_fwd_mb,
        filename="act_total",
    )
    plot_metric(
        metric_name="Activation max tensor (proxy) vs batch size",
        y_label="MB",
        y_getter=lambda r: r.activation_max_single_mb,
        filename="act_max",
    )

    plot_metric(
        metric_name="Latency (ms/image) vs batch size",
        y_label="ms/image",
        y_getter=lambda r: r.latency_ms_per_image,
        filename="latency_ms_per_image",
        skip_all_na=False,
    )
    plot_metric(
        metric_name="Throughput (images/s) vs batch size",
        y_label="images/s",
        y_getter=lambda r: r.throughput_images_s,
        filename="throughput_images_s",
        skip_all_na=False,
    )

    # MACs/FLOPs are per-image and do not depend on batch; plot vs resolution
    if any(r.gmacs_per_image is not None for r in runtime_rows):
        fig, ax = plt.subplots(1, 1, figsize=(6.2, 4.2))
        ax.set_title("Per-image compute vs resolution")
        ax.set_xlabel("resolution")
        ax.set_ylabel("GMACs/img")
        ax.grid(True, alpha=0.3)
        for m in models_sorted:
            # use first batch size entry per resolution
            ys = []
            xs = []
            for res in resolutions_sorted:
                pts = [r for r in runtime_rows if r.model == m and r.resolution == res]
                pts = sorted(pts, key=lambda r: r.batch_size)
                val = next((p.gmacs_per_image for p in pts if p.gmacs_per_image is not None), None)
                if val is None:
                    continue
                xs.append(res)
                ys.append(float(val))
            if xs:
                ax.plot(xs, ys, marker="o", linewidth=1.5, label=m)
        ax.legend(fontsize=8)
        fig.tight_layout()
        out_path = f"{out_prefix}_gmacs_vs_res.png"
        fig.savefig(out_path, dpi=200)
        plt.close(fig)
        print(f"Saved plot: {out_path}")


def _write_csv(path: str, rows: List[dict]) -> None:
    if not rows:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def _print_markdown(static_rows: List[StaticRow], runtime_rows: List[RuntimeRow]) -> None:
    print("\n=== Static metrics ===")
    print("| model | dtype | #params | weights+buffers (MB) |")
    print("|---|---:|---:|---:|")
    for r in static_rows:
        print(f"|{r.model}|{r.dtype}|{r.num_params}|{r.weights_and_buffers_mb:.2f}|")

    print("\n=== Runtime metrics (PEAK during inference) ===")
    print("|m|dev|dt|bs|res|GMAC|GFLOP|ms/img|img/s|pRSS|pGPU|actT|actM|")
    print("|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|")
    for r in runtime_rows:
        gpu = "NA" if r.peak_gpu_allocated_mb is None else f"{r.peak_gpu_allocated_mb:.2f}"
        gmacs = "NA" if r.gmacs_per_image is None else f"{r.gmacs_per_image:.3f}"
        gflops = "NA" if r.gflops_per_image is None else f"{r.gflops_per_image:.3f}"
        print(
            "|"
            + "|".join(
                [
                    r.model,
                    r.device,
                    r.dtype,
                    str(r.batch_size),
                    str(r.resolution),
                    gmacs,
                    gflops,
                    f"{r.latency_ms_per_image:.3f}",
                    f"{r.throughput_images_s:.2f}",
                    f"{r.peak_cpu_rss_mb:.2f}",
                    gpu,
                    f"{r.activation_total_per_fwd_mb:.2f}",
                    f"{r.activation_max_single_mb:.2f}",
                ]
            )
            + "|"
        )

def main() -> None:
    ap = argparse.ArgumentParser(description="Compare CNN memory efficiency (PEAK runtime + MACs/FLOPs)")
    ap.add_argument(
        "--config",
        default=None,
        help="Path to YAML/JSON config file. If provided, values are loaded then overridden by explicit CLI flags.",
    )
    ap.add_argument(
        "--models",
        default="resnet18,mobilenet_v3_small",
        help="Comma-separated: resnet18, mobilenet_v3_small, mobilenet_v3_large",
    )
    ap.add_argument("--device", default="auto", choices=["auto", "cpu", "cuda"])
    ap.add_argument("--dtype", default="fp32", choices=["fp32", "fp16", "bf16"])
    ap.add_argument("--pretrained", action="store_true", help="Load pretrained weights (may download)")
    ap.add_argument("--batch-sizes", default="1-16")
    ap.add_argument("--resolutions", default="128,224")
    ap.add_argument("--warmup", type=int, default=2)
    ap.add_argument("--iters", type=int, default=5)
    ap.add_argument("--out-prefix", default="cnn_compare")
    ap.add_argument("--no-plot", action="store_true", help="Do not generate PNG plots")
    ap.add_argument(
        "--isolate-runs",
        action="store_true",
        help="Rebuild model for every (bs,res) to reduce cross-run effects (slower)",
    )
    defaults = ap.parse_args([])
    args = ap.parse_args()

    if args.config:
        cfg = _read_config(args.config)
        _apply_config_over_args(args, defaults, cfg)

    device = _normalize_device(args.device)
    dtype, dtype_str = _safe_infer_dtype(device, args.dtype)
    if dtype_str != args.dtype:
        print(f"[note] device={device.type} 不支持 {args.dtype}，已退回 {dtype_str}")

    model_names = [m.strip() for m in str(args.models).split(",") if m.strip()]
    batch_sizes = _parse_int_list(args.batch_sizes)
    resolutions = _parse_int_list(args.resolutions)

    # For "一次只使用一个"：要求单次运行只测一个模型（如需多个模型，请多跑几次/多份 config）。
    if len(model_names) != 1:
        raise ValueError(
            f"本次运行检测到 {len(model_names)} 个模型：{model_names}。\n"
            "请在 config/CLI 中只指定一个模型（例如 --models resnet18），一次只跑一个并输出单模型报告。"
        )

    # 用户需求场景一般固定 224；允许多个 resolution，但建议一次一个。
    if len(resolutions) != 1:
        print(f"[note] 当前 resolutions={resolutions}（建议一次只跑一个 resolution，以便报告更清晰）")

    static_rows: List[StaticRow] = []
    runtime_rows: List[RuntimeRow] = []

    # Cache MACs/FLOPs per (model,res) because batch size doesn't change per-image compute.
    macs_cache: Dict[Tuple[str, int], Tuple[Optional[float], Optional[float]]] = {}
    if thop_profile is None:
        print("[note] 未安装 thop，MACs/FLOPs 将显示为 NA。可选安装：pip install thop")

    report_path: Optional[str] = None
    for model_name in model_names:
        # 为避免把 init/下载 等影响混进 runtime，static 与 runtime 分开处理。
        base_model = _load_torchvision_model(model_name, pretrained=args.pretrained)
        num_params, total_numel = _module_params_and_numel_including_buffers(base_model)
        bytes_per_elem = torch.tensor([], dtype=dtype).element_size()
        static_row = StaticRow(
            model=model_name,
            dtype=dtype_str,
            num_params=num_params,
            weights_and_buffers_mb=_mb(int(total_numel * bytes_per_elem)),
        )
        static_rows.append(static_row)

        # Runtime：默认复用同一个模型实例，避免 batch=1-16 时耗时过长
        runtime_model: Optional[torch.nn.Module] = None
        if not args.isolate_runs:
            runtime_model = _load_torchvision_model(model_name, pretrained=args.pretrained)

        for bs in batch_sizes:
            for res in resolutions:
                if (model_name, res) not in macs_cache:
                    macs_cache[(model_name, res)] = _compute_macs_flops_per_image(
                        model_name=model_name,
                        resolution=res,
                        pretrained=args.pretrained,
                    )
                gmacs, gflops = macs_cache[(model_name, res)]

                if args.isolate_runs:
                    model = _load_torchvision_model(model_name, pretrained=args.pretrained)
                    gc.collect()
                    if device.type == "cuda":
                        torch.cuda.empty_cache()
                else:
                    assert runtime_model is not None
                    model = runtime_model

                peak_rss_mb, peak_gpu_mb, act_total_mb, act_max_mb, latency_ms_per_image, throughput_images_s = _run_inference_sweep(
                    model=model,
                    device=device,
                    dtype_str=dtype_str,
                    batch_size=bs,
                    resolution=res,
                    warmup=args.warmup,
                    iters=args.iters,
                )
                runtime_rows.append(
                    RuntimeRow(
                        model=model_name,
                        device=device.type,
                        dtype=dtype_str,
                        batch_size=bs,
                        resolution=res,
                        gmacs_per_image=gmacs,
                        gflops_per_image=gflops,
                        latency_ms_per_image=latency_ms_per_image,
                        throughput_images_s=throughput_images_s,
                        peak_cpu_rss_mb=peak_rss_mb,
                        peak_gpu_allocated_mb=peak_gpu_mb,
                        activation_total_per_fwd_mb=act_total_mb,
                        activation_max_single_mb=act_max_mb,
                    )
                )
                print(
                    f"done: model={model_name} bs={bs} res={res} "
                    f"peakRSS={peak_rss_mb:.2f}MB peakGPU={'NA' if peak_gpu_mb is None else f'{peak_gpu_mb:.2f}MB'}"
                )

        if runtime_model is not None:
            del runtime_model
            gc.collect()

        # 单模型报告：合并 static + runtime + 环境元信息，便于 Excel 一次性查看
        meta = _env_meta(device=device, dtype_str=dtype_str, pretrained=args.pretrained)
        meta["warmup"] = int(args.warmup)
        meta["iters"] = int(args.iters)
        report_path = _write_detailed_report(args.out_prefix, meta, static_row, runtime_rows)

    _print_markdown(static_rows, runtime_rows)

    static_path = f"{args.out_prefix}_static_metrics.csv"
    runtime_path = f"{args.out_prefix}_runtime_metrics.csv"
    _write_csv(static_path, [asdict(r) for r in static_rows])
    _write_csv(runtime_path, [asdict(r) for r in runtime_rows])
    print(f"\nSaved: {static_path}")
    print(f"Saved: {runtime_path}")
    if report_path:
        print(f"Saved: {report_path}")

    if not args.no_plot:
        _plot_runtime(runtime_rows, args.out_prefix)


if __name__ == "__main__":
    main()